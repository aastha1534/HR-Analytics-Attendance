"""
build_excel.py
---------------
Builds HR_Attendance_Analysis.xlsx from the consolidated attendance data:
  - RawData sheet: attendance_monthly.csv (74 employees x 3 months)
  - Dashboard sheet: KPI cards + monthly trend table + top leave-takers
    table, all driven by live SUMIF/AVERAGEIF/COUNTIF formulas.

Run:
    python build_excel.py
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="2E7D5B")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16, color="2E7D5B")
KPI_LABEL_FONT = Font(name=FONT_NAME, size=10, color="595959")
KPI_VALUE_FONT = Font(name=FONT_NAME, bold=True, size=20, color="2E7D5B")
BODY_FONT = Font(name=FONT_NAME, size=10)
thin = Side(style="thin", color="D9D9D9")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

df = pd.read_csv("../data/attendance_monthly.csv")

wb = Workbook()

# ---------------------------------------------------------------- RAW DATA
ws_raw = wb.active
ws_raw.title = "RawData"
for r in dataframe_to_rows(df, index=False, header=True):
    ws_raw.append(r)
for c in range(1, df.shape[1] + 1):
    cell = ws_raw.cell(row=1, column=c)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    ws_raw.column_dimensions[get_column_letter(c)].width = 16
ws_raw.freeze_panes = "A2"
n_rows = df.shape[0] + 1

col_letter = {name: get_column_letter(i + 1) for i, name in enumerate(df.columns)}
def rng(col):
    return f"RawData!${col_letter[col]}$2:${col_letter[col]}${n_rows}"

name_rng = rng("Name")
code_rng = rng("EmployeeCode")
month_rng = rng("Month")
present_rng = rng("Present")
wfh_rng = rng("WorkFromHome")
paidleave_rng = rng("PaidLeave")
sickleave_rng = rng("SickLeave")
lwp_rng = rng("LeaveWithoutPay")

# ---------------------------------------------------------------- DASHBOARD
ws = wb.create_sheet("Dashboard")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
for c in "BCDEFGH":
    ws.column_dimensions[c].width = 18

ws["B2"] = "HR Attendance Dashboard — AtliQ (Apr-Jun 2022)"
ws["B2"].font = TITLE_FONT
ws["B3"] = "Source: RawData sheet (74 employees x 3 months = 222 records) — all figures are live formulas"
ws["B3"].font = Font(name=FONT_NAME, italic=True, size=9, color="808080")

# ---- KPI cards ----
kpis = [
    ("Employees Tracked", f"=SUMPRODUCT(1/COUNTIF({code_rng},{code_rng}))", "#,##0"),
    ("Avg Days Present/Mo", f"=AVERAGE({present_rng})", "0.0"),
    ("Avg WFH Days/Mo", f"=AVERAGE({wfh_rng})", "0.0"),
    ("Total Paid Leave Days", f"=SUM({paidleave_rng})", "#,##0"),
    ("Total Sick Leave Days", f"=SUM({sickleave_rng})", "#,##0"),
    ("Total Unpaid Leave Days", f"=SUM({lwp_rng})", "#,##0"),
]
col_start = 2
for i, (label, formula, fmt) in enumerate(kpis):
    col = col_start + i
    ws.cell(row=5, column=col, value=label).font = KPI_LABEL_FONT
    ws.cell(row=5, column=col).alignment = Alignment(wrap_text=True)
    vcell = ws.cell(row=6, column=col, value=formula)
    vcell.font = KPI_VALUE_FONT
    vcell.number_format = fmt
    for row in (5, 6):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="E9F5EE")
        ws.cell(row=row, column=col).border = BORDER
    ws.row_dimensions[6].height = 30

# ---- Monthly trend table ----
start_row = 9
ws.cell(row=start_row, column=2, value="Monthly Attendance Trend").font = Font(name=FONT_NAME, bold=True, size=12, color="2E7D5B")
headers = ["Month", "Avg Present", "Avg WFH", "Avg Paid Leave", "Avg Sick Leave", "Avg Unpaid Leave"]
for i, h in enumerate(headers):
    cell = ws.cell(row=start_row + 1, column=2 + i, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = BORDER

months = ["Apr 2022", "May 2022", "June 2022"]
for i, m in enumerate(months):
    r = start_row + 2 + i
    ws.cell(row=r, column=2, value=m).font = BODY_FONT
    ws.cell(row=r, column=3, value=f'=AVERAGEIF({month_rng},"{m}",{present_rng})').font = BODY_FONT
    ws.cell(row=r, column=4, value=f'=AVERAGEIF({month_rng},"{m}",{wfh_rng})').font = BODY_FONT
    ws.cell(row=r, column=5, value=f'=AVERAGEIF({month_rng},"{m}",{paidleave_rng})').font = BODY_FONT
    ws.cell(row=r, column=6, value=f'=AVERAGEIF({month_rng},"{m}",{sickleave_rng})').font = BODY_FONT
    ws.cell(row=r, column=7, value=f'=AVERAGEIF({month_rng},"{m}",{lwp_rng})').font = BODY_FONT
    for c in range(2, 8):
        ws.cell(row=r, column=c).border = BORDER
        if c > 2:
            ws.cell(row=r, column=c).number_format = "0.00"

# ---- Top 10 leave-takers table (by total leave across all months) ----
start_row2 = start_row + 2 + len(months) + 2
ws.cell(row=start_row2, column=2, value="Top 10 Employees by Total Leave (Apr-Jun 2022)").font = Font(name=FONT_NAME, bold=True, size=12, color="2E7D5B")
headers2 = ["Rank", "Employee", "Total Paid+Sick+Unpaid Leave"]
for i, h in enumerate(headers2):
    cell = ws.cell(row=start_row2 + 1, column=2 + i, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = BORDER

# Precompute the actual ranking in Python (values only) since a full dynamic
# top-N with formulas would need XLOOKUP/SORT, which are unsupported by
# LibreOffice recalculation in this environment (see xlsx skill notes) -
# instead, each cell here still references RawData live via SUMIF, so if the
# data changes the leave totals recalc even though the top-10 selection is fixed.
leave_totals = (
    df.assign(TotalLeave=df["PaidLeave"] + df["SickLeave"] + df["LeaveWithoutPay"])
      .groupby("Name")["TotalLeave"].sum()
      .sort_values(ascending=False)
      .head(10)
)
for i, (name, _) in enumerate(leave_totals.items()):
    r = start_row2 + 2 + i
    ws.cell(row=r, column=2, value=i + 1).font = BODY_FONT
    ws.cell(row=r, column=3, value=name).font = BODY_FONT
    formula = (f'=SUMIF({name_rng},"{name}",{paidleave_rng})'
               f'+SUMIF({name_rng},"{name}",{sickleave_rng})'
               f'+SUMIF({name_rng},"{name}",{lwp_rng})')
    ws.cell(row=r, column=4, value=formula).font = BODY_FONT
    for c in range(2, 5):
        ws.cell(row=r, column=c).border = BORDER

note_row = start_row2 + 2 + len(leave_totals) + 2
ws.cell(row=note_row, column=2,
        value="Note: KPI cards and trend table use live SUMIF/AVERAGEIF/SUMPRODUCT formulas "
              "referencing RawData. Top-10 ranking order is fixed (computed once from the "
              "source data) but each leave total still recalculates live via SUMIF. "
              "Source: AtliQ attendance workbook, consolidated via /python/consolidate_attendance.py.")
ws.cell(row=note_row, column=2).font = Font(name=FONT_NAME, italic=True, size=8, color="808080")

wb.save("HR_Attendance_Analysis.xlsx")
print("Saved HR_Attendance_Analysis.xlsx")
